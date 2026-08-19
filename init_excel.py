#!/usr/bin/env python3
"""
Script para crear/inicializar el archivo Excel de transacciones
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def crear_excel_base(nombre_archivo="registro_finanzas.xlsx"):
    """Crea un archivo Excel con la estructura necesaria."""
    
    try:
        # Intentar cargar si existe
        wb = openpyxl.load_workbook(nombre_archivo)
        print(f"✅ Archivo {nombre_archivo} ya existe.")
        return
    except:
        pass
    
    # Crear nuevo libro
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    
    # Encabezados
    encabezados = ["Fecha", "Tipo", "Categoría", "Monto", "Descripción", "Origen", "ID Ref"]
    ws.append(encabezados)
    
    # Estilos para el encabezado
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 12  # Fecha
    ws.column_dimensions['B'].width = 10  # Tipo
    ws.column_dimensions['C'].width = 25  # Categoría
    ws.column_dimensions['D'].width = 12  # Monto
    ws.column_dimensions['E'].width = 35  # Descripción
    ws.column_dimensions['F'].width = 15  # Origen
    ws.column_dimensions['G'].width = 15  # ID Ref
    
    # Formatos
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=7):
        row[0].number_format = 'YYYY-MM-DD'  # Fecha
        row[3].number_format = '$#,##0.00'    # Monto
    
    # Agregar algunos datos de ejemplo
    ejemplo_datos = [
        [datetime.now().strftime("%Y-%m-%d"), "Egreso", "Supermercado / Comida", 2500, "Coto - Verduras", "Telegram", "TG-EJEMPLO1"],
        [datetime.now().strftime("%Y-%m-%d"), "Ingreso", "Sueldo / Ingreso", 50000, "Sueldo mensual", "Banco", "BNK-EJEMPLO1"],
    ]
    
    for dato in ejemplo_datos:
        ws.append(dato)
    
    # Guardar
    wb.save(nombre_archivo)
    print(f"✅ Archivo {nombre_archivo} creado exitosamente con estructura base.")
    print(f"📋 Columnas: {', '.join(encabezados)}")
    print(f"📌 Datos de ejemplo agregados")

if __name__ == "__main__":
    crear_excel_base()
