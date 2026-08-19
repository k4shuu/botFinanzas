# ==============================================================================
# BOT DE TELEGRAM DE FINANZAS PERSONALES CON LECTURA DE EMAILS Y REGISTRO EN EXCEL
# ==============================================================================
# Requisitos previos:
# pip install python-telegram-bot openpyxl google-generativeai python-dotenv
#
# Configuración:
# 1. Crea tu bot en Telegram con @BotFather y obtén el TELEGRAM_TOKEN.
# 2. Habilita una "Contraseña de Aplicación" en tu cuenta de Gmail (o IMAP de tu proveedor).
# 3. Obtén tu API key de Google Gemini desde https://makersuite.google.com/app/apikeys
# 4. Guarda el archivo "registro_finanzas.xlsx" en la misma carpeta que este script.
# 5. Crea un archivo .env con tus credenciales (copia .env.example)
# ==============================================================================

import logging
import os
import re
import imaplib
import email
from email.header import decode_header
from datetime import datetime, date
from pathlib import Path
import json
import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import google.generativeai as genai
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# Cargar variables de entorno desde archivo .env
try:
    from dotenv import load_dotenv
    env_path = Path(__file__).parent / ".env"
    load_dotenv(env_path)
    print(f"✅ Variables de entorno cargadas desde: {env_path}")
except ImportError:
    print("⚠️  python-dotenv no instalado, intentando cargar manualmente")
    import sys
    env_path = Path(__file__).parent / ".env"
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    os.environ[key.strip()] = value.strip()
        print(f"✅ Variables cargadas desde: {env_path}")

# ------------------------------------------------------------------------------
# CONFIGURACIÓN GENERAL
# ------------------------------------------------------------------------------
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
MY_TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

# Configuración IMAP (Gmail)
IMAP_SERVER = "imap.gmail.com"
EMAIL_USER = os.getenv("GMAIL_USER", "")
EMAIL_PASS = os.getenv("GMAIL_PASS", "")

# API Key de Gemini
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
if GEMINI_API_KEY:
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        print("✅ Gemini API configurada correctamente")
    except Exception as e:
        print(f"❌ Error configurando Gemini: {e}")
        GEMINI_API_KEY = ""
else:
    print("⚠️  GEMINI_API_KEY no está configurada")

# Google Sheets Configuration
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "")
GOOGLE_CREDENTIALS_FILE = Path(__file__).parent / "google_credentials.json"

# Conectar a Google Sheets
gs_client = None
gs_worksheet = None

if GOOGLE_SHEET_ID and GOOGLE_CREDENTIALS_FILE.exists():
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(str(GOOGLE_CREDENTIALS_FILE), scope)
        gs_client = gspread.authorize(creds)
        gs_spreadsheet = gs_client.open_by_key(GOOGLE_SHEET_ID)
        gs_worksheet = gs_spreadsheet.get_worksheet(0)  # Primera hoja
        print("✅ Google Sheets conectado correctamente")
    except Exception as e:
        print(f"❌ Error conectando a Google Sheets: {e}")
        gs_client = None
        gs_worksheet = None
else:
    print("⚠️  Google Sheets no configurado o archivo de credenciales no encontrado")

EXCEL_FILE = "registro_finanzas.xlsx"

# Reglas simples de Clasificación por Palabras Clave
REGLAS_CATEGORIAS = {
    "coto": ("Supermercado / Comida", "Egreso"),
    "carrefour": ("Supermercado / Comida", "Egreso"),
    "jumbo": ("Supermercado / Comida", "Egreso"),
    "dia": ("Supermercado / Comida", "Egreso"),
    "disco": ("Supermercado / Comida", "Egreso"),
    "pedidosya": ("Salidas / Entretenimiento", "Egreso"),
    "rappi": ("Salidas / Entretenimiento", "Egreso"),
    "glovo": ("Salidas / Entretenimiento", "Egreso"),
    "edea": ("Servicios / Impuestos", "Egreso"),
    "camuzzi": ("Servicios / Impuestos", "Egreso"),
    "telecom": ("Servicios / Impuestos", "Egreso"),
    "personal": ("Servicios / Impuestos", "Egreso"),
    "movistar": ("Servicios / Impuestos", "Egreso"),
    "fibertel": ("Servicios / Impuestos", "Egreso"),
    "transferencia enviada": ("Transferencia Enviada", "Egreso"),
    "transferencia recibida": ("Transferencia Recibida", "Ingreso"),
    "acreditacion": ("Sueldo / Ingreso", "Ingreso"),
    "haberes": ("Sueldo / Ingreso", "Ingreso"),
    "retiro": ("Retiro de Efectivo", "Egreso"),
    "gasolinera": ("Transporte / Combustible", "Egreso"),
    "estacion de servicio": ("Transporte / Combustible", "Egreso"),
    "farmacia": ("Salud", "Egreso"),
    "hospital": ("Salud", "Egreso"),
    "cine": ("Entretenimiento", "Egreso"),
    "restaurante": ("Comidas Afuera", "Egreso"),
    "cafe": ("Comidas Afuera", "Egreso"),
    "clinica": ("Salud", "Egreso"),
    "odontologia": ("Salud", "Egreso"),
}

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)


# ------------------------------------------------------------------------------
# FUNCIONES PARA REGISTRAR EN GOOGLE SHEETS
# ------------------------------------------------------------------------------
def registrar_en_google_sheets(fecha, tipo, categoria, monto, descripcion, origen, id_ref=""):
    """Registra una transacción en Google Sheets (sin hora)."""
    if not gs_worksheet:
        logging.error("Google Sheets no está conectado")
        return False
    
    try:
        nueva_fila = [fecha, tipo, categoria, monto, descripcion, origen, id_ref]
        gs_worksheet.append_row(nueva_fila)
        logging.info(f"Transacción registrada en Google Sheets: {descripcion} (${monto})")
        return True
    except Exception as e:
        logging.error(f"Error al escribir en Google Sheets: {e}")
        return False


def registrar_en_excel(fecha, tipo, categoria, monto, descripcion, origen, id_ref=""):
    """Registra una transacción en Excel (legacy, sin hora)."""
    try:
        if not os.path.exists(EXCEL_FILE):
            # Crear archivo si no existe
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transacciones"
            ws.append(["Fecha", "Tipo", "Categoría", "Monto", "Descripción", "Origen", "ID Ref"])
            wb.save(EXCEL_FILE)
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Transacciones"]

        nueva_fila = [fecha, tipo, categoria, monto, descripcion, origen, id_ref]
        ws.append(nueva_fila)

        # Aplicar formato básico a la nueva fila agregada
        last_row = ws.max_row
        ws.cell(row=last_row, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=last_row, column=4).number_format = "$ #,##0.00"

        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        logging.error(f"Error al escribir en Excel: {e}")
        return False


def clasificar_transaccion(texto):
    """Clasifica una transacción usando palabras clave."""
    texto_lower = texto.lower()
    for kw, (cat, tipo) in REGLAS_CATEGORIAS.items():
        if kw in texto_lower:
            return cat, tipo
    return "Compras Varias", "Egreso"


def extraer_monto(texto):
    """Extrae monto en formato $ o número puro."""
    patrones = [
        r'\$\s*([\d.,]+)',
        r'([\d.,]+)\s*(?:pesos|AR\$|ars)',
        r'^\s*(\d+(?:[.,]\d+)?)\s*',
    ]
    
    for patron in patrones:
        match = re.search(patron, texto, re.IGNORECASE)
        if match:
            monto_raw = match.group(1).replace(".", "").replace(",", ".")
            try:
                return float(monto_raw)
            except ValueError:
                continue
    return None


async def analizar_con_gemini(texto, update):
    """Analiza un mensaje con Gemini para extraer información financiera."""
    if not GEMINI_API_KEY:
        return None
    
    try:
        prompt = f"""Analiza este mensaje financiero en español y extrae la información en formato JSON.
        Responde SOLO con JSON válido, sin comentarios adicionales.
        
        Mensaje: "{texto}"
        
        IMPORTANTE sobre números: los montos en el mensaje están en formato argentino,
        donde el punto (.) separa miles y la coma (,) separa decimales.
        Ejemplos: "4.000,00" = cuatro mil = 4000.0 | "1.500" = mil quinientos = 1500.0 | "3500" = 3500.0
        Devolvé el campo "monto" como número estándar (con punto decimal, sin separador de miles).
        
        Devuelve JSON con estas claves (si no se encontrarn datos, usa null):
        {{
            "monto": <número o null>,
            "tipo": "<Ingreso o Egreso o null>",
            "categoria": "<categoría sugerida o null>",
            "descripcion": "<descripción clara o null>",
            "confianza": <0.0 a 1.0>
        }}
        
        Categorías válidas: Supermercado / Comida, Salidas / Entretenimiento, Servicios / Impuestos, 
        Transferencia Enviada, Transferencia Recibida, Sueldo / Ingreso, Retiro de Efectivo, 
        Transporte / Combustible, Salud, Entretenimiento, Comidas Afuera, Compras Varias, Otros Ingresos"""
        
        model = genai.GenerativeModel('gemini-flash-latest')
        response = model.generate_content(prompt)
        
        import json
        resultado = json.loads(response.text)
        return resultado
    except Exception as e:
        logging.error(f"Error con Gemini: {e}")
        return None


async def parsear_resumen_mercado_pago(texto):
    """Parsea un resumen de Mercado Pago y extrae transacciones individuales."""
    if not GEMINI_API_KEY:
        return []
    
    try:
        prompt = f"""Analiza este resumen de Mercado Pago (puede estar en formato de texto copiado/pegado).
        Ignora COMPLETAMENTE cualquier referencia al "saldo" de la cuenta.
        Extrae SOLO las transacciones individuales (compras, transferencias, pagos, etc.)
        
        IMPORTANTE sobre números: los montos están en formato argentino,
        donde el punto (.) separa miles y la coma (,) separa decimales.
        Ejemplo: "4.000,00" = cuatro mil = 4000.0. Devolvé "monto" como número estándar
        (con punto decimal, sin separador de miles).
        
        Texto: "{texto}"
        
        Responde SOLO con JSON válido, sin comentarios. Devuelve un array de transacciones:
        [
            {{
                "fecha": "YYYY-MM-DD",
                "monto": <número>,
                "descripcion": "<descripción corta>",
                "tipo": "<Ingreso o Egreso>",
                "categoria": "<categoría sugerida>"
            }},
            ...
        ]
        
        Si no encuentras transacciones válidas o solo hay saldo, devuelve: []
        
        Categorías válidas: Supermercado / Comida, Salidas / Entretenimiento, Servicios / Impuestos, 
        Transferencia Enviada, Transferencia Recibida, Sueldo / Ingreso, Retiro de Efectivo, 
        Transporte / Combustible, Salud, Entretenimiento, Comidas Afuera, Compras Varias, Otros Ingresos"""
        
        model = genai.GenerativeModel('gemini-flash-latest')
        response = model.generate_content(prompt)
        
        import json
        transacciones = json.loads(response.text)
        
        if isinstance(transacciones, list):
            logging.info(f"Se extrajeron {len(transacciones)} transacciones de Mercado Pago")
            return transacciones
        return []
    except Exception as e:
        logging.error(f"Error parseando Mercado Pago: {e}")
        return []


def escapar_markdown(texto):
    """Escapa los caracteres especiales del modo Markdown (legacy) de Telegram
    para que texto libre (descripciones, categorías) no rompa el parser.
    Caracteres relevantes en este modo: _ * ` ["""
    if texto is None:
        return ""
    texto = str(texto)
    for caracter in ("_", "*", "`", "["):
        texto = texto.replace(caracter, f"\\{caracter}")
    return texto


def transaccion_ya_existe(fecha, monto, descripcion):
    """Verifica si una transacción ya existe comparando fecha y monto."""
    if not gs_worksheet:
        return False
    
    try:
        rows = gs_worksheet.get_all_values()[1:]  # Sin encabezado
        
        for row in rows:
            if len(row) >= 5:
                fecha_existente = str(row[0]).strip()
                try:
                    monto_existente = float(row[3])
                except (ValueError, IndexError):
                    continue
                
                # Comparar fecha y monto (descripción es flexible)
                if fecha_existente == fecha and abs(monto_existente - float(monto)) < 0.01:
                    return True
        return False
    except Exception as e:
        logging.error(f"Error verificando duplicados: {e}")
        return False


def contiene_referencia_temporal(texto):
    """Chequeo local (sin gastar cuota de Gemini) para saber si vale la pena
    preguntarle a la IA por una fecha. Evita una llamada a Gemini en el caso
    común de transacciones de HOY, que no mencionan ninguna fecha."""
    texto = texto.lower()
    palabras_clave = [
        "ayer", "anteayer", "antes de ayer", "hace ", "semana pasada",
        "mes pasado", "el lunes", "el martes", "el miércoles", "el jueves",
        "el viernes", "el sábado", "el domingo", " de enero", " de febrero",
        " de marzo", " de abril", " de mayo", " de junio", " de julio",
        " de agosto", " de septiembre", " de octubre", " de noviembre",
        " de diciembre", "atrás", "pasado",
    ]
    return any(palabra in texto for palabra in palabras_clave)


async def extraer_fecha_hora_con_gemini(texto):
    """Extrae fecha y hora de un mensaje usando Gemini (para transacciones antiguas)."""
    if not GEMINI_API_KEY:
        return None, None
    
    try:
        prompt = f"""Analiza este mensaje y extrae SOLO la referencia temporal (fecha).
        Responde SOLO con JSON válido, sin comentarios adicionales.
        
        Mensaje: "{texto}"
        Fecha actual: {datetime.now().strftime('%Y-%m-%d')}
        
        Si encuentras referencias a fechas antiguas (ayer, hace 3 días, el 5 de agosto, etc.), 
        calcula la fecha exacta y devuelve:
        {{
            "fecha": "YYYY-MM-DD",
            "tiene_fecha": true
        }}
        
        Si NO hay referencia temporal, devuelve:
        {{
            "fecha": null,
            "tiene_fecha": false
        }}"""
        
        model = genai.GenerativeModel('gemini-flash-latest')
        response = model.generate_content(prompt)
        
        import json
        resultado = json.loads(response.text)
        
        if resultado.get("tiene_fecha"):
            return resultado.get("fecha")
        return None
    except Exception as e:
        logging.warning(f"Error extrayendo fecha con Gemini: {e}")
        return None


# ------------------------------------------------------------------------------
# COMANDOS DEL BOT DE TELEGRAM
# ------------------------------------------------------------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = (
        "👋 **¡Hola! Soy tu Bot de Finanzas Personales.**\n\n"
        "📌 **¿Qué puedo hacer por ti?**\n"
        "1. 💾 Registro automáticamente egresos/ingresos en Google Sheets\n"
        "2. 🤖 Entiendo lenguaje natural gracias a Gemini AI\n"
        "3. 📅 Puedo registrar transacciones antiguas (solo indica la fecha)\n"
        "4. 💬 Registra gastos manuales con ejemplos naturales:\n\n"
        "**Ejemplos de mensajes (HOY):**\n"
        "   • `Gasté 3500 pesos en la verdulería`\n"
        "   • `Compré en Coto $2800`\n"
        "   • `Me llegaron 50000 de freelance`\n"
        "   • `+15000 por vender mis libros`\n"
        "   • `Retiro 5 mil en efectivo`\n"
        "   • `Pague 1200 de luz`\n\n"
        "**Ejemplos de mensajes ANTIGUOS (transacciones olvidadas):**\n"
        "   • `Ayer gasté $2500 en el super`\n"
        "   • `Hace 3 días pagué 8000 de alquiler`\n"
        "   • `El 5 de agosto compré una remera por $1800`\n"
        "   • `La semana pasada cobré 120000 de un proyecto`\n\n"
        "**Comandos disponibles:**\n"
        "   /ayuda - Ver esta ayuda nuevamente\n"
        "   /resumen - Ver resumen del mes\n"
        "   /ultimas - Ver últimas 10 transacciones\n"
        "   /borrar\\_ultimo - Elimina la última transacción ⚠️"
    )
    await update.message.reply_text(msg, parse_mode="Markdown")


async def cmd_ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start(update, context)


async def cmd_ultimas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra las últimas 10 transacciones registradas desde Google Sheets."""
    try:
        if not gs_worksheet:
            await update.message.reply_text("❌ Google Sheets no está conectado.")
            return
        
        # Obtener todas las filas (excepto encabezado)
        rows = gs_worksheet.get_all_values()
        if len(rows) <= 1:
            await update.message.reply_text("📋 No hay transacciones registradas aún.")
            return
        
        ultimas = rows[1:][-10:][::-1]  # Últimas 10, sin encabezado, ordenadas inversas
        msg = "📋 **Últimas 10 Transacciones:**\n\n"
        
        for idx, row in enumerate(ultimas, 1):
            if len(row) >= 7:
                fecha, tipo, categoria, monto, desc, origen, ref = row[:7]
                emoji = "📈" if tipo == "Ingreso" else "📉"
                msg += f"{idx}. {emoji} {fecha} - {tipo}\n"
                msg += f"   💰 ${monto} | {escapar_markdown(categoria)}\n"
                msg += f"   📝 {escapar_markdown(desc)}\n\n"
        
        await update.message.reply_text(msg, parse_mode="Markdown")
    except Exception as e:
        logging.error(f"Error en cmd_ultimas: {e}")
        await update.message.reply_text("❌ Error al leer Google Sheets.")


async def cmd_resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra resumen financiero del mes actual desde Google Sheets."""
    try:
        if not gs_worksheet:
            await update.message.reply_text("❌ Google Sheets no está conectado.")
            return
        
        rows = gs_worksheet.get_all_values()[1:]  # Sin encabezado
        
        mes_actual = datetime.now().strftime("%Y-%m")
        ingresos = 0
        egresos = 0
        categorias = {}
        
        for row in rows:
            if len(row) >= 5 and row[0]:
                fecha = str(row[0])[:7]  # YYYY-MM
                if fecha == mes_actual:
                    try:
                        monto = float(row[3])
                    except (ValueError, IndexError):
                        continue
                    tipo = row[1] if len(row) > 1 else ""
                    categoria = row[2] if len(row) > 2 else ""
                    
                    if tipo == "Ingreso":
                        ingresos += monto
                    else:
                        egresos += monto
                    
                    if categoria not in categorias:
                        categorias[categoria] = 0
                    categorias[categoria] += monto
        
        balance = ingresos - egresos
        emoji_balance = "✅" if balance >= 0 else "⚠️"
        
        msg = f"📊 **Resumen Financiero - {mes_actual}**\n\n"
        msg += f"📈 Ingresos: ${ingresos:,.2f}\n"
        msg += f"📉 Egresos: ${egresos:,.2f}\n"
        msg += f"{emoji_balance} Balance: ${balance:,.2f}\n\n"
        
        if categorias:
            msg += "**Gasto por Categoría:**\n"
            for cat, monto in sorted(categorias.items(), key=lambda x: x[1], reverse=True):
                msg += f"  • {escapar_markdown(cat)}: ${monto:,.2f}\n"
        
        await update.message.reply_text(msg, parse_mode="Markdown")
    except Exception as e:
        logging.error(f"Error en cmd_resumen: {e}")
        await update.message.reply_text("❌ Error al leer el archivo.")


async def cmd_borrar_ultimo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Borra la última transacción registrada (en Google Sheets si está conectado, si no en Excel)."""
    try:
        if gs_worksheet:
            rows = gs_worksheet.get_all_values()
            if len(rows) <= 1:  # Solo encabezado
                await update.message.reply_text("❌ No hay transacciones para borrar.")
                return

            ultima_fila = rows[-1]
            fecha, tipo, categoria, monto_raw = ultima_fila[0], ultima_fila[1], ultima_fila[2], ultima_fila[3]
            desc = ultima_fila[4] if len(ultima_fila) > 4 else ""
            try:
                monto = float(monto_raw)
            except (ValueError, TypeError):
                monto = 0.0

            # La fila de datos N corresponde a la fila N+1 de la hoja (1 = encabezado)
            gs_worksheet.delete_rows(len(rows))
        else:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb["Transacciones"]

            max_row = ws.max_row
            if max_row <= 1:  # Solo encabezado
                await update.message.reply_text("❌ No hay transacciones para borrar.")
                return

            ultima_fila = list(ws[max_row])
            fecha, tipo, categoria, monto, desc, origen, ref = [cell.value for cell in ultima_fila]

            ws.delete_rows(max_row)
            wb.save(EXCEL_FILE)

        emoji = "📈" if tipo == "Ingreso" else "📉"
        msg = (
            f"{emoji} **Transacción Eliminada:**\n"
            f"• **Tipo:** {tipo}\n"
            f"• **Categoría:** {escapar_markdown(categoria)}\n"
            f"• **Monto:** ${monto:,.2f}\n"
            f"• **Detalle:** {escapar_markdown(desc)}\n"
            f"• **Fecha:** {fecha}"
        )
        await update.message.reply_text(msg, parse_mode="Markdown")
        logging.info(f"Transacción borrada: {desc} (${monto})")
    except Exception as e:
        logging.error(f"Error en cmd_borrar_ultimo: {e}")
        await update.message.reply_text("❌ Error al borrar la transacción.")



async def mensaje_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip()
    
    # Verificar si es un resumen de Mercado Pago (con múltiples líneas y transacciones)
    transacciones_mp = []
    if "\n" in texto and len(texto) > 100:
        transacciones_mp = await parsear_resumen_mercado_pago(texto)
    
    # Si se detectaron transacciones de Mercado Pago
    if transacciones_mp:
        registradas = 0
        duplicadas = 0
        
        for trans in transacciones_mp:
            try:
                fecha = trans.get("fecha")
                monto = float(trans.get("monto", 0))
                tipo = trans.get("tipo", "Egreso")
                categoria = trans.get("categoria", "Compras Varias")
                descripcion = trans.get("descripcion", "")[:50]
                
                # Verificar si ya existe
                if transaccion_ya_existe(fecha, monto, descripcion):
                    duplicadas += 1
                    logging.info(f"Transacción duplicada ignorada: {fecha} ${monto}")
                    continue
                
                # Registrar transacción
                ahora = datetime.now()
                éxito = registrar_en_google_sheets(
                    fecha=fecha,
                    tipo=tipo,
                    categoria=categoria,
                    monto=monto,
                    descripcion=descripcion,
                    origen="Mercado Pago (Resumen)",
                    id_ref=f"MP-{ahora.strftime('%d%H%M%S')}"
                )
                
                if not éxito:
                    registrar_en_excel(
                        fecha=fecha,
                        tipo=tipo,
                        categoria=categoria,
                        monto=monto,
                        descripcion=descripcion,
                        origen="Mercado Pago (Resumen)",
                        id_ref=f"MP-{ahora.strftime('%d%H%M%S')}"
                    )
                
                registradas += 1
            except Exception as e:
                logging.error(f"Error registrando transacción MP: {e}")
                continue
        
        # Respuesta al usuario
        msg = f"📊 **Resumen Mercado Pago Procesado:**\n"
        msg += f"✅ {registradas} transacciones registradas\n"
        if duplicadas > 0:
            msg += f"⏭️ {duplicadas} transacciones ignoradas (duplicadas)"
        
        await update.message.reply_text(msg, parse_mode="Markdown")
        return
    
    # TRANSACCIÓN MANUAL INDIVIDUAL
    # Intentar usar Gemini primero si está disponible
    resultado_gemini = None
    if GEMINI_API_KEY:
        resultado_gemini = await analizar_con_gemini(texto, update)
    
    # Si Gemini tuvo éxito y buena confianza, usar esos datos
    if resultado_gemini and resultado_gemini.get("confianza", 0) > 0.7 and resultado_gemini.get("monto"):
        monto = resultado_gemini.get("monto")
        tipo = resultado_gemini.get("tipo", "Egreso")
        categoria = resultado_gemini.get("categoria", "Compras Varias")
        descripcion = resultado_gemini.get("descripcion", texto[:60])
    else:
        # Fallback: usar patrón regex tradicional
        patron = r"^(\+|-)?\s*(\d+(?:[.,]\d+)?)\s+(.+)$"
        match = re.match(patron, texto)
        
        # Si no hay match exacto, intentar extraer solo el monto
        if not match:
            monto = extraer_monto(texto)
            if not monto:
                await update.message.reply_text(
                    "⚠️ No pude entender bien tu mensaje.\n"
                    "Intenta con formatos como:\n"
                    "• `3500 retiro efectivo`\n"
                    "• `Gasté 2800 en el supermercado`\n"
                    "• `+15000 ingreso por venta`"
                )
                return
            
            signo = "+" if "ingreso" in texto.lower() or "cobr" in texto.lower() else ""
            tipo = "Ingreso" if signo == "+" else "Egreso"
            descripcion = texto
            categoria, _ = clasificar_transaccion(texto)
        else:
            signo, monto_str, descripcion = match.groups()
            monto = float(monto_str.replace(",", "."))
            
            if signo == "+":
                tipo = "Ingreso"
                categoria = "Otros Ingresos"
            else:
                tipo = "Egreso"
                categoria, _ = clasificar_transaccion(descripcion)

    # Intentar extraer fecha del mensaje (para transacciones antiguas)
    # Solo si el texto parece mencionar una fecha, para no gastar cuota de Gemini innecesariamente
    fecha_extraida = None
    if contiene_referencia_temporal(texto):
        fecha_extraida = await extraer_fecha_hora_con_gemini(texto)
    
    if fecha_extraida:
        fecha = fecha_extraida
        ahora = datetime.now()  # Para generar ID único
        origen = "Telegram (Manual - Fecha Especificada)"
    else:
        # Si no hay fecha especificada, usar la actual
        ahora = datetime.now()
        fecha = ahora.strftime("%Y-%m-%d")
        origen = "Telegram (Manual)"

    # Verificar si ya existe
    if transaccion_ya_existe(fecha, monto, descripcion[:50]):
        await update.message.reply_text(
            "⚠️ **Transacción Duplicada**\n"
            f"Ya existe un registro con la misma fecha ({fecha}) y monto (${monto:,.2f})."
        )
        return

    # Registrar en Google Sheets (primario) y Excel (respaldo)
    éxito = registrar_en_google_sheets(
        fecha=fecha,
        tipo=tipo,
        categoria=categoria,
        monto=monto,
        descripcion=descripcion[:50],
        origen=origen,
        id_ref=f"TG-{ahora.strftime('%d%H%M%S')}"
    )
    
    # Si Google Sheets falla, intentar con Excel
    if not éxito:
        éxito = registrar_en_excel(
            fecha=fecha,
            tipo=tipo,
            categoria=categoria,
            monto=monto,
            descripcion=descripcion[:50],
            origen=origen,
            id_ref=f"TG-{ahora.strftime('%d%H%M%S')}"
        )

    if éxito:
        emoji = "📈" if tipo == "Ingreso" else "📉"
        confianza_text = f" (IA: {resultado_gemini.get('confianza', 0)*100:.0f}%)" if resultado_gemini else ""
        res_msg = (
            f"{emoji} **¡Transacción Registrada!**{confianza_text}\n"
            f"• **Tipo:** {tipo}\n"
            f"• **Categoría:** {escapar_markdown(categoria)}\n"
            f"• **Monto:** ${monto:,.2f}\n"
            f"• **Detalle:** {escapar_markdown(descripcion)}\n"
            f"• **Fecha:** {fecha}"
        )
        await update.message.reply_text(res_msg, parse_mode="Markdown")
    else:
        await update.message.reply_text("❌ Error al guardar en el archivo Excel.")



# ------------------------------------------------------------------------------
# FUNCIONES AUXILIARES PARA RESUMEN AUTOMÁTICO
# ------------------------------------------------------------------------------
def generar_resumen_financiero(mes=None):
    """Genera resumen financiero desde Google Sheets."""
    try:
        if not gs_worksheet:
            return "❌ Google Sheets no está conectado"
        
        if not mes:
            mes = datetime.now().strftime("%Y-%m")
        
        rows = gs_worksheet.get_all_values()[1:]  # Sin encabezado
        
        ingresos = 0
        egresos = 0
        categorias = {}
        
        for row in rows:
            if len(row) >= 5 and row[0]:
                fecha = str(row[0])[:7]  # YYYY-MM
                if fecha == mes:
                    try:
                        monto = float(row[3])
                    except (ValueError, IndexError):
                        continue
                    tipo = row[1] if len(row) > 1 else ""
                    categoria = row[2] if len(row) > 2 else ""
                    
                    if tipo == "Ingreso":
                        ingresos += monto
                    else:
                        egresos += monto
                    
                    if categoria not in categorias:
                        categorias[categoria] = 0
                    categorias[categoria] += monto
        
        balance = ingresos - egresos
        emoji_balance = "✅" if balance >= 0 else "⚠️"
        
        msg = f"📊 **RESUMEN MENSUAL - {mes}** (Google Sheets)\n\n"
        msg += f"📈 Ingresos Total: ${ingresos:,.2f}\n"
        msg += f"📉 Egresos Total: ${egresos:,.2f}\n"
        msg += f"{emoji_balance} Balance Neto: ${balance:,.2f}\n\n"
        
        if categorias:
            msg += "**Top 5 Categorías de Gasto:**\n"
            top_categorias = sorted(categorias.items(), key=lambda x: x[1], reverse=True)[:5]
            for idx, (cat, monto) in enumerate(top_categorias, 1):
                porcentaje = (monto / egresos * 100) if egresos > 0 else 0
                msg += f"{idx}. {cat}: ${monto:,.2f} ({porcentaje:.1f}%)\n"
        
        return msg
    except Exception as e:
        logging.error(f"Error generando resumen: {e}")
        return f"❌ Error al generar resumen: {e}"


async def enviar_resumen_automatico(app_context):
    """Envía resumen automático a fin de mes."""
    try:
        chat_id = int(MY_TELEGRAM_CHAT_ID) if MY_TELEGRAM_CHAT_ID else None
        if not chat_id:
            logging.warning("TELEGRAM_CHAT_ID no configurado para resumen automático")
            return
        
        mes = datetime.now().strftime("%Y-%m")
        resumen = generar_resumen_financiero(mes)
        
        logging.info(f"Enviando resumen automático para {mes} a chat {chat_id}")
        print(f"📊 Enviando resumen automático a fin de mes...")
    except Exception as e:
        logging.error(f"Error enviando resumen automático: {e}")


# ------------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------------
if __name__ == "__main__":
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    # Comandos
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("ayuda", cmd_ayuda))
    app.add_handler(CommandHandler("ultimas", cmd_ultimas))
    app.add_handler(CommandHandler("resumen", cmd_resumen))
    app.add_handler(CommandHandler("borrar_ultimo", cmd_borrar_ultimo))
    
    # Manejador de mensajes de texto (para transacciones manuales)
    app.add_handler(MessageHandler(filters.TEXT & (~filters.COMMAND), mensaje_manual))

    # Configurar scheduler para resumen automático a fin de mes
    scheduler = BackgroundScheduler()
    
    # Resumen a las 23:59 del día 31 (para meses con 31 días)
    scheduler.add_job(
        enviar_resumen_automatico,
        CronTrigger(day=31, hour=23, minute=59),
        args=[app],
        id="resumen_fin_mes_31",
        name="Resumen Automático Día 31"
    )
    
    # Resumen a las 23:59 del día 30 (para febrero y otros meses)
    scheduler.add_job(
        enviar_resumen_automatico,
        CronTrigger(day=30, hour=23, minute=59),
        args=[app],
        id="resumen_fin_mes_30",
        name="Resumen Automático Día 30"
    )
    
    # Resumen a las 23:59 del día 28 (para febrero)
    scheduler.add_job(
        enviar_resumen_automatico,
        CronTrigger(day=28, hour=23, minute=59),
        args=[app],
        id="resumen_fin_mes_28",
        name="Resumen Automático Día 28"
    )
    
    scheduler.start()

    print("🤖 Bot iniciado correctamente y listo para operar...")
    print("🔗 Conectado a Telegram:", "✅" if TELEGRAM_TOKEN else "❌")
    print("🤖 Gemini API:", "✅" if GEMINI_API_KEY else "⚠️ No configurado")
    print("📧 Email:", "✅" if EMAIL_USER else "❌")
    print("📅 Scheduler:", "✅ Resumen automático a fin de mes programado")
    print()
    app.run_polling()